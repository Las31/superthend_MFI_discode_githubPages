"""
NASDAQ 100 SuperTrend (MFI Dynamic Multiplier) Scanner
=======================================================
策略 C：MFI 動態調整 ATR 乘數
  dynamic_mult = base_mult * (1 - (MFI - 50) / 100 * weight)
  - MFI 高（量大）→ 乘數縮小 → 上下軌收緊 → 訊號更靈敏
  - MFI 低（量縮）→ 乘數放寬 → 上下軌擴張 → 過濾假突破

Weight 自動回測：對每支股票在 [0.1, 0.3, 0.5, 0.7] 中選勝率最高的值

安裝套件：
  pip install finvizfinance yfinance pandas openpyxl requests

設定 Webhook：
  set DISCORD_WEBHOOK_URL=https://discord.com/api/webhooks/xxxxx

執行：
  python supertrend_scanner.py
  python supertrend_scanner.py --atr 10 --mult 3 --days 3 --mfi 14
"""

import sys, argparse, datetime, warnings, os
warnings.filterwarnings("ignore")

# ── 設定 Discord Webhook ──────────────────────────────────────
DISCORD_WEBHOOK_URL = os.environ.get("DISCORD_WEBHOOK_URL", "https://discord.com/api/webhooks/1483485512407711967/p27W2VM5ueLR5nSLFon6Yr4gRAf6U9C5AEYfq-fkbuxb_4upTlzt05546wcUpVWFlJ8K")
# 或直接填入：
# DISCORD_WEBHOOK_URL = "https://discord.com/api/webhooks/..."

# ── 參數解析 ──────────────────────────────────────────────────
parser = argparse.ArgumentParser()
parser.add_argument("--atr",    type=int,   default=10,    help="ATR 週期")
parser.add_argument("--mult",   type=float, default=3.0,   help="ATR 基礎乘數")
parser.add_argument("--mfi",    type=int,   default=14,    help="MFI 週期")
parser.add_argument("--days",   type=int,   default=3,     help="觀察翻多天數")
parser.add_argument("--period", type=str,   default="6mo", help="yfinance 資料區間")
parser.add_argument("--no-discord", action="store_true")
args = parser.parse_args()

ATR_PERIOD   = args.atr
BASE_MULT    = args.mult
MFI_PERIOD   = args.mfi
FLIP_DAYS    = args.days
DATA_PERIOD  = args.period
SKIP_DISCORD = args.no_discord

# weight 候選值（回測自動選最佳）
WEIGHT_CANDIDATES = [0.1, 0.3, 0.5, 0.7]

# ── 套件檢查 ──────────────────────────────────────────────────
missing = []
for pkg in ["finvizfinance", "yfinance", "pandas", "openpyxl", "requests"]:
    try:
        __import__(pkg.replace("-", "_"))
    except ImportError:
        missing.append(pkg)
if missing:
    print("[ERROR] Missing packages:")
    print(f"        pip install {' '.join(missing)}")
    sys.exit(1)

import pandas as pd
import yfinance as yf


# ── yfinance 強化下載（retry + backoff + timeout）────────────
import time as _time
import random as _random

_YF_MAX_RETRIES  = 2
_YF_BASE_BACKOFF = 3      # 3s -> skip
_YF_JITTER       = (1, 3)

def yf_download(ticker: str, **kwargs) -> "pd.DataFrame":
    """
    yf.download() with exponential backoff retry.
    Passes all kwargs directly to yf.download().
    """
    kwargs.setdefault("progress", False)
    kwargs.setdefault("auto_adjust", True)

    last_err = None
    for attempt in range(1, _YF_MAX_RETRIES + 1):
        try:
            df = yf.download(ticker, **kwargs)
            if df is not None and not df.empty:
                return df
            raise ValueError(f"Empty DataFrame returned for {ticker}")
        except Exception as e:
            last_err = e
            if attempt < _YF_MAX_RETRIES:
                wait = _YF_BASE_BACKOFF * (2 ** (attempt - 1)) + _random.uniform(*_YF_JITTER)
                print(f"  [yfinance] {ticker} attempt {attempt}/{_YF_MAX_RETRIES} failed: {e}")
                print(f"  [yfinance] Retrying in {wait:.1f}s ...")
                _time.sleep(wait)

    raise RuntimeError(
        f"[yfinance] {ticker}: all {_YF_MAX_RETRIES} attempts failed. "
        f"Last error: {last_err}"
    )

import requests
from finvizfinance.screener.overview import Overview
from get_tickers import get_nasdaq100_tickers as _fetch_tickers


# ══════════════════════════════════════════════════════════════
# 指標計算
# ══════════════════════════════════════════════════════════════

def compute_mfi(df: pd.DataFrame, period: int) -> pd.Series:
    """
    Money Flow Index (MFI)
    = 100 - 100 / (1 + PMF / NMF)
    Typical Price = (H + L + C) / 3
    Raw Money Flow = TP × Volume
    """
    tp  = (df["High"] + df["Low"] + df["Close"]) / 3
    rmf = tp * df["Volume"]

    pos_mf = rmf.where(tp > tp.shift(1), 0.0)
    neg_mf = rmf.where(tp < tp.shift(1), 0.0)

    pmf = pos_mf.rolling(period).sum()
    nmf = neg_mf.rolling(period).sum()

    mfi = 100 - (100 / (1 + pmf / nmf.replace(0, float("nan"))))
    return mfi.fillna(50)  # 無量時填 50（中性）


def compute_supertrend_dynamic(df: pd.DataFrame,
                                atr_period: int,
                                base_mult: float,
                                mfi_period: int,
                                weight: float) -> pd.DataFrame:
    """
    策略 C：MFI 動態調整乘數的 SuperTrend
    dynamic_mult = base_mult * (1 - (MFI - 50) / 100 * weight)
    """
    high  = df["High"]
    low   = df["Low"]
    close = df["Close"]

    # ATR (Wilder smoothing)
    tr = pd.concat([
        high - low,
        (high - close.shift(1)).abs(),
        (low  - close.shift(1)).abs(),
    ], axis=1).max(axis=1)
    atr = tr.ewm(alpha=1 / atr_period, adjust=False).mean()

    # MFI
    mfi = compute_mfi(df, mfi_period)

    # 動態乘數：MFI=80 → mult 縮小；MFI=20 → mult 放大
    dynamic_mult = base_mult * (1 - (mfi - 50) / 100 * weight)
    # 防止乘數過小或負值
    dynamic_mult = dynamic_mult.clip(lower=base_mult * 0.3)

    # SuperTrend 計算
    hl2         = (high + low) / 2
    basic_upper = hl2 + dynamic_mult * atr
    basic_lower = hl2 - dynamic_mult * atr

    upper      = basic_upper.copy()
    lower      = basic_lower.copy()
    direction  = pd.Series(index=df.index, dtype=int)
    supertrend = pd.Series(index=df.index, dtype=float)

    for i in range(1, len(df)):
        upper.iloc[i] = (
            basic_upper.iloc[i]
            if basic_upper.iloc[i] < upper.iloc[i-1] or close.iloc[i-1] > upper.iloc[i-1]
            else upper.iloc[i-1]
        )
        lower.iloc[i] = (
            basic_lower.iloc[i]
            if basic_lower.iloc[i] > lower.iloc[i-1] or close.iloc[i-1] < lower.iloc[i-1]
            else lower.iloc[i-1]
        )
        if close.iloc[i] > upper.iloc[i-1]:
            direction.iloc[i] = 1
        elif close.iloc[i] < lower.iloc[i-1]:
            direction.iloc[i] = -1
        else:
            direction.iloc[i] = direction.iloc[i-1]

        supertrend.iloc[i] = lower.iloc[i] if direction.iloc[i] == 1 else upper.iloc[i]

    df = df.copy()
    df["ATR"]          = atr
    df["MFI"]          = mfi
    df["DynMult"]      = dynamic_mult
    df["ST_Upper"]     = upper
    df["ST_Lower"]     = lower
    df["SuperTrend"]   = supertrend
    df["ST_Direction"] = direction
    return df


def backtest_weight(df: pd.DataFrame, atr_period: int,
                    base_mult: float, mfi_period: int,
                    candidates: list) -> float:
    """
    對同一段歷史資料，用不同 weight 各跑一次 SuperTrend，
    計算翻多後 5 日平均報酬率，選勝率最高的 weight。
    """
    best_weight = candidates[0]
    best_score  = -999.0
    close       = df["Close"]

    for w in candidates:
        result = compute_supertrend_dynamic(df, atr_period, base_mult, mfi_period, w)
        result["ST_Prev"] = result["ST_Direction"].shift(1)
        flips = result[(result["ST_Direction"] == 1) & (result["ST_Prev"] == -1)]

        if len(flips) < 2:
            continue

        # 翻多後 5 日報酬
        returns = []
        for idx in flips.index:
            loc = close.index.get_loc(idx)
            if loc + 5 < len(close):
                ret = (close.iloc[loc + 5] - close.iloc[loc]) / close.iloc[loc]
                returns.append(ret)

        if returns:
            score = sum(1 for r in returns if r > 0) / len(returns)  # 勝率
            if score > best_score:
                best_score  = score
                best_weight = w

    return best_weight


# ══════════════════════════════════════════════════════════════
# Step 1: Finviz 股票清單
# ══════════════════════════════════════════════════════════════
def get_tickers():
    return _fetch_tickers(verbose=True)


def get_nasdaq100_tickers():
    return _fetch_tickers(verbose=True)

# ══════════════════════════════════════════════════════════════
# Step 2+3: 掃描翻多
# ══════════════════════════════════════════════════════════════
def scan_flip_bullish(tickers: list) -> pd.DataFrame:
    print(f"\n[Step 2+3] Scanning {len(tickers)} stocks "
          f"| ATR={ATR_PERIOD}, BaseMult={BASE_MULT}, MFI={MFI_PERIOD}, "
          f"FlipDays={FLIP_DAYS}, AutoWeight...")

    today   = pd.Timestamp.today().normalize()
    cutoff  = today - pd.Timedelta(days=FLIP_DAYS)
    results       = []
    errors        = []   # tickers that failed after retries
    n_skip_empty  = 0    # skipped due to empty / too-short data
    n_total       = len(tickers)

    for i, ticker in enumerate(tickers, 1):
        try:
            raw = yf_download(ticker, period=DATA_PERIOD, interval="1d")
            if raw.empty or len(raw) < max(ATR_PERIOD, MFI_PERIOD) + 20:
                n_skip_empty += 1
                continue
            if isinstance(raw.columns, pd.MultiIndex):
                raw.columns = raw.columns.get_level_values(0)

            # 自動回測選最佳 weight
            best_w = backtest_weight(raw, ATR_PERIOD, BASE_MULT, MFI_PERIOD, WEIGHT_CANDIDATES)

            # 用最佳 weight 計算最終指標
            df = compute_supertrend_dynamic(raw, ATR_PERIOD, BASE_MULT, MFI_PERIOD, best_w)
            df = df.dropna(subset=["ST_Direction"])

            df["ST_Prev"] = df["ST_Direction"].shift(1)
            flips = df[(df["ST_Direction"] == 1) & (df["ST_Prev"] == -1)]
            recent = flips[flips.index >= cutoff]
            if recent.empty:
                continue

            last_flip  = recent.iloc[-1]
            flip_date  = last_flip.name.date()
            last_close = float(df["Close"].iloc[-1])
            flip_price = float(last_flip["Close"])
            change_pct = (last_close - flip_price) / flip_price * 100
            mfi_at_flip = round(float(last_flip["MFI"]), 1)
            dyn_mult    = round(float(last_flip["DynMult"]), 3)

            # MFI 評級（輔助參考）
            if mfi_at_flip >= 70:
                mfi_grade = "Strong"
            elif mfi_at_flip >= 50:
                mfi_grade = "Normal"
            else:
                mfi_grade = "Weak"

            results.append({
                "Ticker"      : ticker,
                "Flip Date"   : flip_date,
                "Flip Price"  : round(flip_price, 2),
                "Last Price"  : round(last_close, 2),
                "Change %"    : round(change_pct, 2),
                "MFI"         : mfi_at_flip,
                "MFI Grade"   : mfi_grade,
                "Dyn Mult"    : dyn_mult,
                "Best Weight" : best_w,
                "SuperTrend"  : round(float(last_flip["SuperTrend"]), 2),
                "ATR"         : round(float(last_flip["ATR"]), 2),
            })

            print(f"  [{i:3d}/{len(tickers)}] {ticker:<6} FLIP {flip_date} "
                  f"| MFI={mfi_at_flip:5.1f} ({mfi_grade:<6}) "
                  f"| DynMult={dyn_mult:.3f} | Weight={best_w}")

        except Exception as e:
            errors.append(ticker)
            print(f"  [{i:3d}/{len(tickers)}] {ticker:<6} SKIP (2 retries exhausted): {e}")

    # ── Download statistics ──────────────────────────
    n_failed  = len(errors)
    n_success = n_total - n_failed - n_skip_empty
    fail_rate = n_failed / n_total * 100 if n_total else 0
    skip_rate = n_skip_empty / n_total * 100 if n_total else 0
    print(f"\n  [Download Stats]")
    print(f"    Total    : {n_total}")
    print(f"    Success  : {n_success}")
    print(f"    Skipped  : {n_skip_empty}  ({skip_rate:.1f}%  no/short data)")
    print(f"    Failed   : {n_failed}  ({fail_rate:.1f}%  exhausted retries)")
    if errors:
        print(f"    Tickers  : {errors}")

    result_df = pd.DataFrame(results)
    if not result_df.empty:
        result_df = result_df.sort_values(
            ["MFI Grade", "Change %"],
            ascending=[True, False]   # Strong 排最前，同級按漲幅排
        ).reset_index(drop=True)
    return result_df


# ══════════════════════════════════════════════════════════════
# Step 4: 輸出
# ══════════════════════════════════════════════════════════════
def get_output_path(filename: str) -> str:
    script_dir = os.path.dirname(os.path.abspath(__file__))
    ts   = datetime.datetime.now().strftime("%Y%m%d_%H%M")
    base, ext = os.path.splitext(filename)
    return os.path.join(script_dir, f"{base}_{ts}{ext}")


def print_results(df: pd.DataFrame):
    print("\n" + "=" * 80)
    print(f"  Bullish Flip | ATR={ATR_PERIOD}, BaseMult={BASE_MULT}, "
          f"MFI={MFI_PERIOD}, AutoWeight, Days={FLIP_DAYS}")
    print("=" * 80)
    if df.empty:
        print("  No bullish flip found.")
        return
    cols = ["Ticker","Flip Date","Flip Price","Last Price","Change %",
            "MFI","MFI Grade","Dyn Mult","Best Weight"]
    print(df[cols].to_string(index=False))
    print(f"\n  Total: {len(df)} | Strong: {(df['MFI Grade']=='Strong').sum()} "
          f"| Normal: {(df['MFI Grade']=='Normal').sum()} "
          f"| Weak: {(df['MFI Grade']=='Weak').sum()}")


def save_csv(df: pd.DataFrame, filename="supertrend_mfi_dynamic.csv"):
    path = get_output_path(filename)
    try:
        df.to_csv(path, index=False, encoding="utf-8-sig")
        print(f"[SAVED] CSV -> {path}")
    except PermissionError:
        desktop = os.path.join(os.path.expanduser("~"), "Desktop", os.path.basename(path))
        df.to_csv(desktop, index=False, encoding="utf-8-sig")
        print(f"[SAVED] CSV -> {desktop} (fallback: Desktop)")


def save_excel(df: pd.DataFrame, filename="supertrend_mfi_dynamic.xlsx"):
    path = get_output_path(filename)
    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Bullish Flip"

        # 顏色定義
        FILLS = {
            "header" : PatternFill("solid", fgColor="1F4E79"),
            "Strong" : PatternFill("solid", fgColor="C6EFCE"),  # 深綠
            "Normal" : PatternFill("solid", fgColor="FFEB9C"),  # 黃
            "Weak"   : PatternFill("solid", fgColor="FFCCCC"),  # 粉紅
        }
        thin = Border(*[Side(style="thin")] * 0,
                      left=Side(style="thin"), right=Side(style="thin"),
                      top=Side(style="thin"),  bottom=Side(style="thin"))

        headers = list(df.columns)
        for ci, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=ci, value=h)
            c.fill = FILLS["header"]
            c.font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
            c.alignment = Alignment(horizontal="center")
            c.border = thin
        ws.row_dimensions[1].height = 20

        for ri, row in df.iterrows():
            grade = row.get("MFI Grade", "Normal")
            fill  = FILLS.get(grade, FILLS["Normal"])
            for ci, val in enumerate(row, 1):
                c = ws.cell(row=ri + 2, column=ci, value=val)
                c.fill = fill
                c.font = Font(name="Arial", size=10)
                c.border = thin
                c.alignment = Alignment(horizontal="center")
                col = headers[ci - 1]
                if col in ("Flip Price","Last Price","SuperTrend","ATR"):
                    c.number_format = "#,##0.00"
                elif col == "Change %":
                    c.number_format = '#,##0.00"%"'
                elif col == "MFI":
                    c.number_format = "#,##0.0"
                elif col in ("Dyn Mult","Best Weight"):
                    c.number_format = "0.000"

        for ci, col in enumerate(headers, 1):
            mx = max(len(str(col)),
                     *(len(str(df.iloc[r][col])) for r in range(len(df)))) + 4
            ws.column_dimensions[get_column_letter(ci)].width = min(mx, 22)
        ws.freeze_panes = "A2"

        # Parameters sheet
        ws2 = wb.create_sheet("Parameters")
        for r, (k, v) in enumerate([
            ("Scan Date",       datetime.datetime.now().strftime("%Y-%m-%d %H:%M")),
            ("ATR Period",      ATR_PERIOD),
            ("Base Multiplier", BASE_MULT),
            ("MFI Period",      MFI_PERIOD),
            ("Weight Method",   f"Auto backtest from {WEIGHT_CANDIDATES}"),
            ("Flip Days",       FLIP_DAYS),
            ("Data Period",     DATA_PERIOD),
            ("Universe",        "NASDAQ 100 (via Finviz)"),
            ("Formula",         "dynamic_mult = base_mult*(1-(MFI-50)/100*weight)"),
        ], 1):
            ws2.cell(row=r, column=1, value=k).font = Font(bold=True, name="Arial")
            ws2.cell(row=r, column=2, value=v).font = Font(name="Arial")
        ws2.column_dimensions["A"].width = 20
        ws2.column_dimensions["B"].width = 45

        wb.save(path)
        print(f"[SAVED] Excel -> {path}")
    except Exception as e:
        print(f"[WARN] Excel failed: {e}")
        save_csv(df, filename)


# ══════════════════════════════════════════════════════════════
# Step 5: Discord Webhook
# ══════════════════════════════════════════════════════════════
# MFI Grade -> Discord embed color & label
GRADE_META = {
    "Strong": {"color": 0x2ECC71, "label": "[Strong] Volume-confirmed flip"},
    "Normal": {"color": 0xF1C40F, "label": "[Normal] Moderate volume flip"},
    "Weak"  : {"color": 0x95A5A6, "label": "[Weak]   Low volume flip"},
}

def send_discord(df: pd.DataFrame, webhook_url: str):
    if not webhook_url:
        print("[SKIP] DISCORD_WEBHOOK_URL not set.")
        print("       set DISCORD_WEBHOOK_URL=https://discord.com/api/webhooks/...")
        return

    scan_time = datetime.datetime.now().strftime("%Y-%m-%d %H:%M")
    total     = len(df)
    strong    = (df["MFI Grade"] == "Strong").sum() if not df.empty else 0
    normal    = (df["MFI Grade"] == "Normal").sum() if not df.empty else 0
    weak      = (df["MFI Grade"] == "Weak").sum()   if not df.empty else 0

    # Summary embed
    summary = {
        "title": "NASDAQ 100  SuperTrend x MFI Dynamic Scanner",
        "description": (
            f"Bullish flips within **{FLIP_DAYS} days** | "
            f"ATR={ATR_PERIOD}  BaseMult={BASE_MULT}  MFI={MFI_PERIOD}  AutoWeight\n\n"
            f"Total: **{total}** stocks  |  "
            f"Strong: **{strong}**  Normal: **{normal}**  Weak: **{weak}**"
        ),
        "color": 0x1F4E79,
        "footer": {"text": f"Scan time: {scan_time}  |  Source: Finviz + yfinance"},
    }
    if df.empty:
        summary["description"] += "\n\n_No bullish flip found._"
    _post_webhook(webhook_url, {"embeds": [summary]}, "summary")
    if df.empty:
        return

    # 分批，每則最多 10 個 embed
    BATCH = 10
    for start in range(0, total, BATCH):
        batch  = df.iloc[start: start + BATCH]
        embeds = []
        for _, row in batch.iterrows():
            meta   = GRADE_META.get(row["MFI Grade"], GRADE_META["Normal"])
            change = row["Change %"]
            arrow  = "+" if change >= 0 else ""
            embed  = {
                "title" : f"{row['Ticker']}  {meta['label']}",
                "color" : meta["color"],
                "fields": [
                    {"name": "Flip Date",    "value": str(row["Flip Date"]),      "inline": True},
                    {"name": "Flip Price",   "value": f"${row['Flip Price']}",    "inline": True},
                    {"name": "Last Price",   "value": f"${row['Last Price']}",    "inline": True},
                    {"name": "Change %",     "value": f"{arrow}{change:.2f}%",   "inline": True},
                    {"name": "MFI",          "value": f"{row['MFI']:.1f}",        "inline": True},
                    {"name": "Dynamic Mult", "value": f"{row['Dyn Mult']:.3f}",   "inline": True},
                    {"name": "Best Weight",  "value": str(row["Best Weight"]),    "inline": True},
                    {"name": "SuperTrend",   "value": f"${row['SuperTrend']}",    "inline": True},
                    {"name": "ATR",          "value": str(row["ATR"]),            "inline": True},
                ],
            }
            embeds.append(embed)
        _post_webhook(webhook_url, {"embeds": embeds}, f"batch {start//BATCH+1}")

    print("[DISCORD] All messages sent.")


def _post_webhook(url, payload, label=""):
    try:
        r = requests.post(url, json=payload, timeout=10)
        status = "OK" if r.status_code in (200, 204) else f"HTTP {r.status_code}"
        print(f"[DISCORD] ({label}) -> {status}")
    except requests.exceptions.ConnectionError:
        print(f"[DISCORD] ERROR: Cannot connect. Check URL or network.")
    except Exception as e:
        print(f"[DISCORD] ERROR ({label}): {e}")


# ══════════════════════════════════════════════════════════════
# Main
# ══════════════════════════════════════════════════════════════
def main():
    print("=" * 80)
    print("  NASDAQ 100  SuperTrend x MFI Dynamic Multiplier  Scanner")
    print(f"  ATR={ATR_PERIOD}  BaseMult={BASE_MULT}  MFI={MFI_PERIOD}  "
          f"AutoWeight={WEIGHT_CANDIDATES}  Days={FLIP_DAYS}")
    print("=" * 80)

    tickers   = get_nasdaq100_tickers()
    result_df = scan_flip_bullish(tickers)

    print_results(result_df)

    if not result_df.empty:
        save_csv(result_df)
        save_excel(result_df)

    if not SKIP_DISCORD:
        print("\n[Step 5] Sending to Discord...")
        send_discord(result_df, DISCORD_WEBHOOK_URL)
    else:
        print("\n[Step 5] Discord skipped.")


if __name__ == "__main__":
    main()